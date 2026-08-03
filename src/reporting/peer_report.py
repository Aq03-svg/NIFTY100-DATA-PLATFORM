"""
Peer Comparison Excel Report

Exports ranked peer comparison data to Excel.
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.screener.scoring import calculate_quality_score


INPUT_FILE = Path("data/processed/financial_ratios.csv")
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "peer_comparison.xlsx"


def load_data():
    """
    Load financial ratios.
    """

    return pd.read_csv(INPUT_FILE)


def prepare_report(df):
    """
    Prepare report data.
    """

    report = df.copy()

    report = report.sort_values(
        by="composite_quality_score",
        ascending=False,
    )

    report["rank"] = range(
        1,
        len(report) + 1,
    )

    columns = [
        "rank",
        "company",
        "return_on_equity",
        "return_on_capital_employed",
        "net_profit_margin",
        "debt_to_equity",
        "free_cash_flow",
        "revenue_cagr",
        "composite_quality_score",
    ]

    return report[columns]


def export_excel(df):
    """
    Export peer comparison report.
    """

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "All Companies"

    gold_fill = PatternFill(
        fill_type="solid",
        start_color="FFD700",
        end_color="FFD700",
    )

    headers = list(df.columns)

    # Write header row
    for column_number, header in enumerate(
        headers,
        start=1,
    ):

        cell = sheet.cell(
            row=1,
            column=column_number,
        )

        cell.value = header
        cell.font = Font(bold=True)

    # Write data rows
    for row in df.itertuples(index=False):

        sheet.append(row)

    # Freeze header row
    sheet.freeze_panes = "A2"

    # Auto-adjust column widths
    for column_cells in sheet.columns:

        length = max(
            len(str(cell.value))
            if cell.value is not None else 0
            for cell in column_cells
        )

        sheet.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = length + 3

    # Highlight top-ranked company
    if sheet.max_row >= 2:
        for cell in sheet[2]:
            cell.fill = gold_fill

    workbook.save(OUTPUT_FILE)

    print(f"Report saved to: {OUTPUT_FILE}")


def main():

    df = load_data()

    df = calculate_quality_score(df)

    report = prepare_report(df)

    export_excel(report)


if __name__ == "__main__":
    main()